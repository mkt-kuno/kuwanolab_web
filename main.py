import os
import io
import gc
import json
import typing
import fastapi
import openpyxl
import resource
import datetime
import jpholiday
import mysql.connector
import mysql.connector.pooling
from dateutil.relativedelta import relativedelta

gc.enable()
app = fastapi.FastAPI()

DB_HOST = "kuwanolabserver.iis.u-tokyo.ac.jp"
DB_NAME = "felica_db"
DB_USER = "bw304"
DB_PASS = "UpomPmBu"
DB_PORT = 3306
pool = mysql.connector.pooling.MySQLConnectionPool(user=DB_USER, password=DB_PASS, database=DB_NAME, host=DB_HOST, port=DB_PORT, pool_size=16)

def get_mariadb_con():
    return pool.get_connection()

def get_datetime_list(name, ymd_from, ymd_to):
    ret = {}

    if name is None or ymd_from is None or ymd_to is None:
        ret["result"] = "error"
        return ret

    con = get_mariadb_con()
    cursor = con.cursor()
    query = "SELECT idm_datetime.`datetime` FROM idm_datetime JOIN idm_name ON (idm_datetime.idm = idm_name.idm) WHERE idm_datetime.`datetime` BETWEEN '"
    query += ymd_from + "' AND '" + ymd_to + "' AND idm_name.name = '" + name +"';"
    cursor.execute(query)
    list = []
    for item in cursor.fetchall():
        #if name not in list:
        list.append(item[0])
    ret["result"] = "success"
    ret["name"] = name
    ret["from"] = ymd_from
    ret["to"] = ymd_to
    ret["datetime"] = list

    cursor.close()
    del cursor
    con.close()
    del con

    return ret

def get_name_list():
    con = get_mariadb_con()
    cursor = con.cursor()
    query = "SELECT name FROM idm_name WHERE enable = 1 ORDER BY priority;"
    cursor.execute(query)
    ret = []
    for item in cursor.fetchall():
        name = item[0]
        if name not in ret:
            ret.append(name)

    cursor.close()
    del cursor
    con.close()
    del con

    return ret

def get_john_doe_list():
    con = get_mariadb_con()
    cursor = con.cursor()
    query  = "SELECT idm_datetime.*,idm_name.name "
    query += "FROM idm_datetime LEFT JOIN idm_name ON (idm_datetime.idm = idm_name.idm) "
    query += "WHERE idm_name.name is NULL "
    query += "ORDER BY idm_datetime.`datetime` DESC LIMIT 10;"
    cursor.execute(query)
    ret = {}
    list = []
    for item in cursor.fetchall():
        idm = item[0]
        datetime = item[1]
        if idm not in list:
            list.append(idm)
            ret[idm] = str(datetime)

    cursor.close()
    del cursor
    con.close()
    del con

    return ret

@app.get('/api/today_list')
def api_today_list():
    return fastapi.responses.JSONResponse(get_today_list())

def get_today_list():
    ret = {}

    today = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + relativedelta(days=1)

    for name in get_name_list():
        dict = get_datetime_list(name, today.strftime("%Y%m%d"), tomorrow.strftime("%Y%m%d"))
        (enter, exit) = find_enter_exit_time(today, dict["datetime"])
        list = []
        for dt in dict["datetime"]:
            list.append(dt.strftime('%Y-%m-%d %H:%M:%S'))
        ret[dict["name"]]  = list
    return ret

def get_holiday_class(day):
    ret = None
    if jpholiday.is_holiday(day) or day.weekday() == 6:
        ret = "holiday_red"
    elif day.weekday() == 5:
        ret = "holiday_blue"
    return ret

def is_same_day(day_a, day_b):
    ret = False
    if day_a.year == day_b.year and day_a.month == day_b.month and day_a.day == day_b.day:
        ret = True
    return ret

def is_today(day):
    ret = False
    today = datetime.datetime.today()
    if today.year == day.year and today.month == day.month and today.day == day.day:
        ret = True
    return ret

def is_thismonth(day):
    ret = False
    today = datetime.datetime.today()
    if today.year == day.year and today.month == day.month:
        ret = True
    return ret

def get_datetime_from_to(yearmonth = None):
    if yearmonth is not None:
        dt_yearmonth = datetime.datetime.strptime(yearmonth, '%Y-%m')
        thismonth = dt_yearmonth.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        thismonth = datetime.datetime.today().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    nextmonth = thismonth + relativedelta(months=1)
    prevmonth = thismonth + relativedelta(months=-1)
    return (prevmonth, thismonth, nextmonth)

def find_enter_exit_time(day, list):
    enter = exit = None;
    # 最初と最後のタッチを判別する
    for item in list:
        if is_same_day(item, day):
            if enter is None:
                enter = item
            elif enter > item:
                enter = item
            if exit is None:
                exit = item
            elif exit < item:
                exit = item
    if enter is not None and exit is not None:
        diff_min = (exit - enter).total_seconds() / 60
        if diff_min < 5:
            exit = None
    return (enter, exit)

def calender_2d_list(yearmonth, just_mark = False):
    ret = {}

    (prevmonth, thismonth, nextmonth) = get_datetime_from_to(yearmonth)

    temp = []
    day = thismonth
    # Table Class
    while day < nextmonth:
        holiday_class = get_holiday_class(day)
        today = is_today(day)
        if holiday_class is not None and today is True:
            temp.append('today' + holiday_class)
        elif today is True:
            temp.append('today')
        elif holiday_class is not None:
            temp.append(holiday_class)
        else:
            temp.append('')
        day += datetime.timedelta(days=1)
    ret["Class"] = temp

    # Table Header
    temp = []
    day = thismonth
    while day < nextmonth:
        temp.append(day.strftime("%Y/%m/%d %a"))
        day += datetime.timedelta(days=1)
    ret["Name"] = temp

    # Table Body
    for name in get_name_list():
        temp = []

        day = thismonth
        dict = get_datetime_list(name, thismonth.strftime("%Y%m%d"), nextmonth.strftime("%Y%m%d"))
        list = dict["datetime"]
        while day < nextmonth:
            (enter, exit) = find_enter_exit_time(day, list)
            if enter is not None:
                if just_mark:
                    temp.append("o")
                elif exit is None:
                    temp.append(enter.strftime("%H:%M"))
                else:
                    temp.append(enter.strftime("%H:%M") + "-" + exit.strftime("%H:%M"))
            else:
                temp.append("")
            day += datetime.timedelta(days=1)
        ret[name] = temp

    return ret

def calender_html_body(yearmonth):
    (prevmonth, thismonth, nextmonth) = get_datetime_from_to(yearmonth)
    return calender_html_body_from_to(thismonth, nextmonth)

def calender_html_body_from_to(day_from, day_to, show_month = False):
    ret = ""
    ret += "\t"+'<div>'+"\r\n"
    ret += "\t\t"+'<center><table class="calender_table">'+"\r\n"
    ret += "\t\t\t"+'<tr><th>Name</th>'

    day = day_from
    # Table Header
    while day < day_to:
        holiday_class = get_holiday_class(day)
        today = is_today(day)

        if holiday_class is not None and today is True:
            ret += '<th class="today %s">' % holiday_class
        elif today is True:
            ret += '<th class="today">'
        elif holiday_class is not None:
            ret += '<th class="%s">' % holiday_class
        else:
            ret += '<th>'
        if show_month:
            ret += day.strftime("%b<br>%d<br>%a")
        else:
            ret += day.strftime("%d<br>%a")
        ret += "</th>"
        day += datetime.timedelta(days=1)
    ret += '</tr>'+"\r\n"

    # Table Body
    for name in get_name_list():
        ret += "\t\t\t"+'<tr><td>'+ name + '</td>'
        day = day_from
        dict = get_datetime_list(name, day_from.strftime("%Y%m%d"), day_to.strftime("%Y%m%d"))
        list = dict["datetime"]
        while day < day_to:
            (enter, exit) = find_enter_exit_time(day, list)
            if is_today(day):
                ret += '<td class="today">'
            else:
                ret += "<td>"
            if enter is not None:
                ret += enter.strftime("%H:%M")
            if exit is not None:
                ret += '<br>' + exit.strftime("%H:%M")
            ret += '</td>'
            day += datetime.timedelta(days=1)
        ret += '</tr>' + "\r\n"

    ret += "\t\t"+'</table></center>'+"\r\n"
    ret += "\t"+'</div>'+"\r\n"
    return ret

@app.get('/api/name_list')
def api_name_list():
    return fastapi.responses.JSONResponse(get_name_list())

@app.get('/api/john_doe_list')
def api_john_doe_list():
    return fastapi.responses.JSONResponse(get_john_doe_list())

@app.get('/api/get_datetime_list')
def api_get_datetime_list(
    name: typing.Optional[str] = None,
    ymd_from: typing.Optional[str] = None,
    ymd_to: typing.Optional[str] = None):

    ret = get_datetime_list(name, ymd_from, ymd_to)
    list = ret['datetime']

    override = []
    for item in list:
        override.append(item.strftime('%Y-%m-%d %H:%M:%S'))
    ret['datetime'] = override

    return fastapi.responses.JSONResponse(ret)

@app.get('/favicon.ico')
def favicon():
    return fastapi.responses.FileResponse('favicon.ico')

@app.get('/dark-theme.css')
def favicon():
    return fastapi.responses.FileResponse('dark-theme.css')

@app.get('/calender')
def app_calprev_next_button_htmlender(yearmonth: typing.Optional[str] = None):
    if yearmonth is None:
        yearmonth = datetime.datetime.today().strftime('%Y-%m')

    ret = ""
    ret += header_html(False)
    ret += calender_html_body(yearmonth)
    ret += footer_html()

    return fastapi.responses.HTMLResponse(ret)

@app.get("/weekly_calender")
def app_get_weekly_calender():
    today = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    day_from = today - relativedelta(days=12)
    day_to = today + relativedelta(days=2)

    ret = ""
    ret += header_html(True)
    ret += calender_html_body_from_to(day_from, day_to, True)
    ret += footer_html()

    return fastapi.responses.HTMLResponse(ret)

@app.get('/api/calender_json')
def app_get_calender_json(yearmonth: typing.Optional[str] = None):
    if yearmonth is None:
        yearmonth = datetime.datetime.today().strftime('%Y-%m')

    ret = calender_2d_list(yearmonth, False)

    return fastapi.responses.JSONResponse(ret)

def prev_next_button_html(yearmonth):
    (prevmonth, thismonth, nextmonth) = get_datetime_from_to(yearmonth)
    this = thismonth.strftime('%Y-%m')
    prev = prevmonth.strftime('%Y-%m')
    next = nextmonth.strftime('%Y-%m')

    ret  = "\t" + '<div>' + "\r\n"
    ret += "\t\t" + '<form action="" method="get">' + '<b><font size="4"> ' + this + ' </font></b>' + "\r\n"
    ret += "\t\t\t" + '<button type="submit" style="height:24px">This Month</button>' + "\r\n"
    ret += "\t\t\t" + '<button type="submit" name="yearmonth" style="height:24px" value="' + prev + '">Prev Month</button>' + "\r\n"
    if is_thismonth(thismonth) is not True:
        ret+= "\t\t\t" + '<button type="submit" name="yearmonth" style="height:24px;" value="' + next + '">Next Month</button>' + "\r\n"
    ret += "\t\t" + '</form>' + "\r\n"
    ret += "\t" + '</div>' + "\r\n"
    return ret

def output_button_html(yearmonth):
    ret = "\t" + '<div>' + "\r\n"
    ret += "\t\t" + '<form action="" method="get">' + "\r\n"
    ret += "\t\t\t" + '<input type="hidden" name="yearmonth" value="%s">' % yearmonth + "\r\n"
    ret += "\t\t\t" + '<button type="submit" name="mode"  style="height:24px" value="csv">CSV Download</button>' + "\r\n"
    ret += "\t\t\t" + '<button type="submit" name="mode"  style="height:24px" value="xlsx">Excel Download</button>' + "\r\n"
    ret += "\t\t" + '</form>' + "\r\n"
    ret += "\t" + '</div>' + "\r\n"
    return ret

def header_html(refresh = True):
    ret  = '<!DOCTYPE html>' + "\r\n"
    ret += '<html>' + "\r\n"
    ## --> Start Header
    ret += '<head>' + "\r\n"
    ret += "\t" + '<meta charset="UTF-8">' + "\r\n"
    ret += "\t" + '<meta name="viewport" content="width=device-width">' + "\r\n"
    ret += "\t" + '<link rel="stylesheet" href="dark-theme.css"/>' + "\r\n"
    ret += "\t" + '<link rel="icon" href="favicon.ico"/>' + "\r\n"
    ret += "\t" + '<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate" />' + "\r\n"
    ret += "\t" + '<meta http-equiv="Pragma" content="no-cache" />' + "\r\n"
    ret += "\t" + '<meta http-equiv="Expires" content="0" />' + "\r\n"
    if refresh is True:
        ret += "\t" + '<meta http-equiv="refresh" content="3">' + "\r\n"
    ret += "\t" + '<title>Bw304 Entry/Exit Record</title>' + "\r\n"
    ret += '</head>' + "\r\n"
    ## <-- End Header
    ret += '<body width="1200px">' + "\r\n"
    return ret

def footer_html():
    ret  = '</body>' + "\r\n"
    ret += '</html>' + "\r\n"
    return ret

def mode_html(yearmonth):
    (prevmonth, thismonth, nextmonth) = get_datetime_from_to(yearmonth)
    refresh = True
    if is_thismonth(thismonth) is not True:
        refresh = False

    ret = ""
    ret += header_html(refresh)
    ret += prev_next_button_html(yearmonth)
    ret += calender_html_body(yearmonth)
    ret += output_button_html(yearmonth)
    ret += footer_html()

    return fastapi.responses.HTMLResponse(ret)

def mode_csv(yearmonth):
    dict = calender_2d_list(yearmonth, True)
    headers = {}
    headers["Content-Description"] = "File Transfer"
    headers["Content-Disposition"] = "attachment; filename=%s.csv" % yearmonth
    headers["Content-Type"] = "application/octet-stream"
    headers["Content-Transfer-Encoding"] = "binary"
    headers["Cache-Control"] = "must-revalidate, post-check=0, pre-check=0"
    headers["Expires"] = "0"

    ret = ""

    for (key,val) in dict.items():
        ret += key
        for v in val:
            ret += ',' + v
        ret += '\r\n'

    return fastapi.responses.Response(content=ret, headers=headers)

def mode_excel(yearmonth):
    dict = calender_2d_list(yearmonth, False)
    headers = {}
    headers["Content-Description"] = "File Transfer"
    headers["Content-Disposition"] = "attachment; filename=%s.xlsx" % yearmonth
    headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    headers["Content-Transfer-Encoding"] = "binary"
    headers["Cache-Control"] = "must-revalidate, post-check=0, pre-check=0"
    headers["Expires"] = "0"

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    row = 1
    for (key,val) in dict.items():
        col = 1
        ws.cell(row, col).value = key.replace(" ", '\n')
        ws.cell(row, col).alignment = openpyxl.styles.Alignment(wrapText=True)
        col += 1
        for v in val:
            ws.cell(row, col).value = v.replace("-", '\n')
            ws.cell(row, col).alignment = openpyxl.styles.Alignment(wrapText=True)
            col += 1
        row += 1

    bio = io.BytesIO()
    wb.save(bio)
    ret = bio.getvalue()

    return fastapi.responses.Response(content=ret, headers=headers)


@app.get("/")
def root(
    yearmonth: typing.Optional[str] = None,
    mode: typing.Optional[str] = None):

    gc.collect()

    if yearmonth is None:
        yearmonth = datetime.datetime.today().strftime('%Y-%m')

    if mode is None:
        return mode_html(yearmonth)
    elif mode == "csv":
        return mode_csv(yearmonth)
    elif mode == "xlsx":
        return mode_excel(yearmonth)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=28080)
